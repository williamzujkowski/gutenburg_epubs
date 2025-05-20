"""Export functionality for book metadata."""

import csv
import json
import logging
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .database import BookDatabase

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """Supported export formats."""
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    MARKDOWN = "markdown"


class BookExporter:
    """Export book metadata to various formats."""
    
    def __init__(self, db_path: str = "gutenberg_books.db"):
        """Initialize exporter.
        
        Args:
            db_path: Path to database
        """
        self.db = BookDatabase(db_path)
    
    def export_books(
        self,
        format: ExportFormat,
        output_path: Path,
        limit: Optional[int] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> bool:
        """Export books to specified format.
        
        Args:
            format: Export format
            output_path: Output file path
            limit: Maximum number of books to export
            filters: Optional filters (language, subject, etc.)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Get books from database
            books = self._get_books(limit, filters)
            
            if not books:
                logger.warning("No books found to export")
                return False
            
            # Export based on format
            if format == ExportFormat.CSV:
                return self._export_csv(books, output_path)
            elif format == ExportFormat.JSON:
                return self._export_json(books, output_path)
            elif format == ExportFormat.EXCEL:
                return self._export_excel(books, output_path)
            elif format == ExportFormat.MARKDOWN:
                return self._export_markdown(books, output_path)
            else:
                logger.error(f"Unknown export format: {format}")
                return False
                
        except Exception as e:
            logger.error(f"Export error: {e}")
            return False
    
    def _get_books(
        self,
        limit: Optional[int] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """Get books from database with optional filters.
        
        Args:
            limit: Maximum number of books
            filters: Optional filters
            
        Returns:
            List of book dictionaries
        """
        with self.db._get_connection() as conn:
            cursor = conn.cursor()
            
            # Build query
            query = """
                SELECT 
                    b.book_id,
                    b.title,
                    b.language,
                    b.download_count,
                    b.metadata,
                    GROUP_CONCAT(DISTINCT a.name) as authors,
                    GROUP_CONCAT(DISTINCT s.subject_name) as subjects,
                    GROUP_CONCAT(DISTINCT f.format_type) as formats
                FROM books b
                LEFT JOIN book_authors ba ON b.book_id = ba.book_id
                LEFT JOIN authors a ON ba.author_id = a.author_id
                LEFT JOIN book_subjects bs ON b.book_id = bs.book_id
                LEFT JOIN subjects s ON bs.subject_id = s.subject_id
                LEFT JOIN formats f ON b.book_id = f.book_id
                WHERE 1=1
            """
            
            params = []
            
            # Apply filters
            if filters:
                if 'language' in filters:
                    query += " AND b.language = ?"
                    params.append(filters['language'])
                
                if 'min_downloads' in filters:
                    query += " AND b.download_count >= ?"
                    params.append(filters['min_downloads'])
                
                if 'subject' in filters:
                    query += " AND s.subject_name LIKE ?"
                    params.append(f"%{filters['subject']}%")
            
            query += " GROUP BY b.book_id ORDER BY b.download_count DESC"
            
            if limit:
                query += f" LIMIT {limit}"
            
            cursor.execute(query, params)
            
            # Convert to list of dictionaries
            books = []
            for row in cursor.fetchall():
                book = dict(row)
                
                # Parse metadata
                if book['metadata']:
                    metadata = json.loads(book['metadata'])
                    book.update(metadata)
                
                # Split concatenated fields
                book['authors'] = book['authors'].split(',') if book['authors'] else []
                book['subjects'] = book['subjects'].split(',') if book['subjects'] else []
                book['formats'] = book['formats'].split(',') if book['formats'] else []
                
                books.append(book)
            
            return books
    
    def _export_csv(self, books: List[Dict[str, Any]], output_path: Path) -> bool:
        """Export books to CSV format.
        
        Args:
            books: List of book dictionaries
            output_path: Output file path
            
        Returns:
            True if successful
        """
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                # Determine fieldnames
                fieldnames = [
                    'book_id', 'title', 'authors', 'language', 
                    'download_count', 'subjects', 'formats'
                ]
                
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for book in books:
                    row = {
                        'book_id': book['book_id'],
                        'title': book['title'],
                        'authors': '; '.join(book['authors']),
                        'language': book['language'],
                        'download_count': book['download_count'],
                        'subjects': '; '.join(book['subjects']),
                        'formats': '; '.join(book['formats'])
                    }
                    writer.writerow(row)
            
            logger.info(f"Exported {len(books)} books to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"CSV export error: {e}")
            return False
    
    def _export_json(self, books: List[Dict[str, Any]], output_path: Path) -> bool:
        """Export books to JSON format.
        
        Args:
            books: List of book dictionaries
            output_path: Output file path
            
        Returns:
            True if successful
        """
        try:
            # Clean up books for JSON
            clean_books = []
            for book in books:
                clean_book = {
                    'book_id': book['book_id'],
                    'title': book['title'],
                    'authors': book['authors'],
                    'language': book['language'],
                    'download_count': book['download_count'],
                    'subjects': book['subjects'],
                    'formats': book['formats']
                }
                clean_books.append(clean_book)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(clean_books, f, indent=2, ensure_ascii=False)
            
            logger.info(f"Exported {len(books)} books to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"JSON export error: {e}")
            return False
    
    def _export_excel(self, books: List[Dict[str, Any]], output_path: Path) -> bool:
        """Export books to Excel format.
        
        Args:
            books: List of book dictionaries
            output_path: Output file path
            
        Returns:
            True if successful
        """
        try:
            # Convert to DataFrame
            df_data = []
            for book in books:
                df_data.append({
                    'Book ID': book['book_id'],
                    'Title': book['title'],
                    'Authors': '; '.join(book['authors']),
                    'Language': book['language'],
                    'Downloads': book['download_count'],
                    'Subjects': '; '.join(book['subjects']),
                    'Formats': '; '.join(book['formats'])
                })
            
            df = pd.DataFrame(df_data)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Books"
            
            # Add headers with formatting
            headers = list(df.columns)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092",
                    end_color="366092",
                    fill_type="solid"
                )
            
            # Add data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Exported {len(books)} books to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            return False
    
    def _export_markdown(self, books: List[Dict[str, Any]], output_path: Path) -> bool:
        """Export books to Markdown format.
        
        Args:
            books: List of book dictionaries
            output_path: Output file path
            
        Returns:
            True if successful
        """
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("# Project Gutenberg Books\n\n")
                f.write(f"Exported {len(books)} books\n\n")
                
                # Create table
                f.write("| Book ID | Title | Authors | Language | Downloads |\n")
                f.write("|---------|-------|---------|----------|----------|\n")
                
                for book in books[:100]:  # Limit to 100 for readability
                    title = book['title'][:50] + "..." if len(book['title']) > 50 else book['title']
                    authors = '; '.join(book['authors'])[:30] + "..." if len('; '.join(book['authors'])) > 30 else '; '.join(book['authors'])
                    
                    f.write(f"| {book['book_id']} | {title} | {authors} | {book['language']} | {book['download_count']:,} |\n")
                
                if len(books) > 100:
                    f.write(f"\n... and {len(books) - 100} more books\n")
            
            logger.info(f"Exported {len(books)} books to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Markdown export error: {e}")
            return False
    
    def export_search_results(
        self,
        search_results: List[Dict[str, Any]],
        format: ExportFormat,
        output_path: Path
    ) -> bool:
        """Export search results to specified format.
        
        Args:
            search_results: List of search result dictionaries
            format: Export format
            output_path: Output file path
            
        Returns:
            True if successful
        """
        # Convert search results to book format
        books = []
        for result in search_results:
            book = {
                'book_id': result.get('book_id', ''),
                'title': result.get('title', ''),
                'authors': [author.get('name', '') for author in result.get('authors', [])],
                'language': result.get('language', ''),
                'download_count': result.get('download_count', 0),
                'subjects': result.get('subjects', []),
                'formats': list(result.get('formats', {}).keys())
            }
            books.append(book)
        
        # Use regular export methods
        if format == ExportFormat.CSV:
            return self._export_csv(books, output_path)
        elif format == ExportFormat.JSON:
            return self._export_json(books, output_path)
        elif format == ExportFormat.EXCEL:
            return self._export_excel(books, output_path)
        elif format == ExportFormat.MARKDOWN:
            return self._export_markdown(books, output_path)
        else:
            logger.error(f"Unknown export format: {format}")
            return False